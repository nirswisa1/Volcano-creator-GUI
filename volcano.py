from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt
from tkinter import *
import math

excelN = ''; title = ''
fcmin = ''; fcmax = ''
pval = ''; spec = ''
i = 0; i2 = 0
selection = ''


def myClick():
    global excelN; global title
    global fcmin; global fcmax
    global pval; global i
    global i2; global spec
    global selection
    excelN = excelE.get()
    title = titleE.get()
    fcmax = eval(fcE.get())
    fcmin = eval(fcE1.get())
    pval = -math.log10(eval(pvalE.get()))
    i = var.get()
    i2 = var2.get()
    spec = specE.get()
    selection = varR.get()
    root.destroy()


# GUI
root = Tk()
root.title('Nir.S - Volcano')

# labels
labelexcel = Label(root, text='Excel name: ')
labeltitle = Label(root, text='title: ')
labelfc = Label(root, text='Fold change: ( x>= | x<= )  ')
labelpval = Label(root, text='p-Val:')

labelexcel.grid(row=0, column=0, sticky='W')
labeltitle.grid(row=1, column=0, sticky='W')
labelfc.grid(row=2, column=0, sticky='W')
labelpval.grid(row=3, column=0, sticky='W')
# slots
excelE = Entry(root, width=25, borderwidth=2)
titleE = Entry(root, width=25, borderwidth=2)
fcE = Entry(root, width=12, borderwidth=2)
fcE1 = Entry(root, width=12, borderwidth=2)
pvalE = Entry(root, width=25, borderwidth=2)
specE = Entry(root, width=25, borderwidth=2)

excelE.grid(row=0, column=1)
titleE.grid(row=1, column=1)
fcE.grid(row=2, column=1, sticky='W')
fcE1.grid(row=2, column=1, sticky='E')
pvalE.grid(row=3, column=1)
specE.grid(row=6, column=1)
# Check box
var = IntVar()
c = Checkbutton(root, text="All labels", variable=var)
c.grid(row=8, column=0, sticky='W')

var2 = IntVar()
c2 = Checkbutton(root, text="Specified labels(optional)", variable=var2)
c2.grid(row=6, column=0, sticky='W')
# Separator
varR = StringVar()
R1 = Radiobutton(root, text="Space", variable=varR, value=' ')
R1.grid(row=7, column=1, sticky='W')
R2 = Radiobutton(root, text="Comma ( , ) ", variable=varR, value=', ')
R2.grid(row=7, column=1, sticky='E')

# Button
b = Button(root, text="Run", width=20, command=myClick)
b.grid(row=8, column=1)

root.mainloop()

fc = [math.log2(fcmax), math.log2(fcmin)]

# excel

wb = load_workbook(excelN + '.xlsx')
ws = wb.active

greens = []
reds = []

for row in range(2, len(ws["F"]) + 1):

    try:
        ws["F" + str(row)].value = -(math.log10(ws["B" + str(row)].value))

        if ws["F" + str(row)].value >= pval and ws["D" + str(row)].value >= fc[0]:
            reds.append([ws["A" + str(row)].value, ws["F" + str(row)].value, ws["D" + str(row)].value])
        elif ws["F" + str(row)].value >= pval and ws["D" + str(row)].value <= fc[1]:
            greens.append([ws["A" + str(row)].value, ws["F" + str(row)].value, ws["D" + str(row)].value])

    except:
        ws["F" + str(row)].value = 'None'

ws["F1"].value = '-Log p-Val'
wb.save(excelN + '.xlsx')

pval = [i.value for i in ws["F"]]
fold = [i.value for i in ws["D"]]

plt.style.use('seaborn')
new_data = {'pval': list(pval[1:]), 'fold': list(fold[1:])}
df = pd.DataFrame.from_dict(new_data)
plt.scatter(list(fold[1:]), list(pval[1:]), color='silver', s=25)
plt.scatter([x[2] for x in reds], [x[1] for x in reds], color='orangered', s=25)
plt.scatter([x[2] for x in greens], [x[1] for x in greens], color='limegreen', s=25)

if i:
    for p in greens:
        plt.text(x=p[2], y=p[1] + 0.05, s=p[0], size=8)
    for p in reds:
        plt.text(x=p[2], y=p[1], s=p[0], size=8)
elif i2:
    for s in spec.split(selection):
        for p in greens:
            if p[0] == s:
                plt.text(x=p[2], y=p[1] + 0.05, s=p[0], size=8)
        for p2 in reds:
            if p2[0] == s:
                plt.text(x=p2[2], y=p2[1] + 0.05, s=p2[0], size=8)

plt.title(title)
plt.xlabel('-Log2 Difference')
plt.ylabel('-Log10 p-Val')
plt.tight_layout()
ax = plt.gca()
ax.legend(['Non-sig', 'up', 'down'])
ax.set_facecolor('whitesmoke')

plt.show()
